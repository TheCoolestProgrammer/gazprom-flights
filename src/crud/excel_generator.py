"""
Module for generating Excel files for flight manifests and passenger lists.
"""
import os
from datetime import datetime
from io import BytesIO
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from src.models.flights import Flight
from src.models.passenger_flight import PassengerFlight
from src.models.cargo import Cargo


def get_template_path(filename: str) -> str:
    """Get the full path to a template file in the docs folder."""
    return os.path.join(str(Path(__file__).parent.parent.parent), "docs", filename)


# def safe_set_cell_value(ws, cell_ref, value):
#     """
#     Safely set cell value, handling merged cells.
#     If cell is part of a merged range, writes to the top-left cell of the range.
#     """
#     try:
#         ws[cell_ref].value = value
#     except AttributeError:
#         # Cell is a MergedCell, find the merged range and write to top-left cell
#         from openpyxl.utils import get_column_letter, column_index_from_string
        
#         # Parse cell reference
#         col_idx = column_index_from_string(cell_ref.rstrip('0123456789'))
#         row_idx = int(cell_ref[len(cell_ref.rstrip('0123456789')):])
        
#         # Find which merged range this cell belongs to
#         for merged_range in ws.merged_cells.ranges:
#             if cell_ref in merged_range:
#                 # Found the merged range, write to top-left cell
#                 top_left = merged_range.start_cell.coordinate
#                 ws[top_left].value = value
#                 return
#         # If not in any merged range, just write anyway
#         ws[cell_ref].value = value


def generate_passenger_manifest_excel(flight: Flight, session: Session) -> bytes:
    """
    Generate the main passenger manifest Excel file from the template.
    Fills in flight information and passenger list.
    
    Template: ОБРАЗЕЦ!!Список пассажиров.xlsx - "Список" sheet
    """
    template_path = get_template_path("ОБРАЗЕЦ!!Список пассажиров.xlsx")
    
    # Load template
    wb = load_workbook(template_path)
    ws = wb["Список"]
    
    # Get passengers for this flight
    passenger_flights = session.query(PassengerFlight).filter(
        PassengerFlight.flight_id == flight.id
    ).all()
    
    # Fill in flight information
    # Row 9: "К заявке №" - Insert flight ID
    # Row 9, Column D contains "К заявке №"
    ws["D9"] = f"К заявке № {flight.id}"
    
    # Row 9: "от" - Insert flight date
    ws["E9"] = f"от {flight.departure_date.strftime('%d.%m.%Y')}"
    
    # Row 10: Aircraft type
    ws["D10"] = f"ВС    {flight.aircraft_type_rel.name}                                           RA"
    
    # Row 11: Flight number
    ws["D11"] = f"№ рейса ГЗП   {flight.flight_number}"
    
    # Row 11: KVS (Pilot name)
    if flight.pilot:
        ws["G11"] = "КВС: "+flight.pilot.name
    
    # Row 12 already has headers, start filling from row 14 (or 15?)
    # Let me check: Row 12 has headers, so passengers start from row 13
    
    # Add passengers starting from row 14
    start_row = 14
    for idx, pf in enumerate(passenger_flights, 1):
        passenger = pf.passengers
        
        row = start_row + idx - 1
        
        # Column C: № пп (number)
        ws[f"C{row}"] = idx
        
        # Column D: Фамилия, Имя, Отчество
        ws[f"D{row}"] = passenger.fullname
        
        # Column G: Номер документа (Passport)
        ws[f"G{row}"] = str(passenger.passport)
        
        # Column H: Дата и время вылета
        departure_time = flight.departure_time.strftime('%H:%M') if flight.departure_time else ""
        ws[f"H{row}"] = f"{flight.departure_date.strftime('%d.%m.%Y')} {departure_time}"
        
        # Column L: Маршрут - из пункта
        if flight.route:
            parts = flight.route.split("-")
            if len(parts) > 0:
                ws[f"L{row}"] = parts[0]  # From point
            if len(parts) > 1:
                ws[f"M{row}"] = parts[1]  # To point

    ws = wb["Груз"]
    cargo_items = session.query(Cargo).filter(Cargo.flight_id == flight.id).all()
    for idx, cargo in enumerate(cargo_items, 1):
        row = 7 + idx
        ws[f"A{row}"] = idx
        ws[f"H{row}"] = cargo.name
        ws[f"DN{row}"] = cargo.packaging_type.value
        ws[f"GP{row}"] = cargo.places_count
        ws[f"HF{row}"] = cargo.weight
        ws[f"FL{row}"] = cargo.flight_from.name
        ws[f"GA{row}"] = cargo.flight_to.name
        # ws[f"G{row}"] = cargo.weight
        # ws[f"H{row}"] = cargo.places_count
        # ws[f"L{row}"] = cargo.flight_from.name if cargo.flight_from else ""
        # ws[f"M{row}"] = cargo.flight_to.name if cargo.flight_to else ""
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()


def generate_ticket_issuance_list_excel(flight: Flight, session: Session) -> bytes:
    """
    Generate the ticket issuance list Excel file.
    
    Template: Список_пассажиров_для_оформления_авиабилетов.xlsx
    """
    template_path = get_template_path("Список_пассажиров_для_оформления_авиабилетов.xlsx")
    
    # Load template
    wb = load_workbook(template_path)
    ws = wb["список пассажиров"]
    
    # Get passengers for this flight
    passenger_flights = session.query(PassengerFlight).filter(
        PassengerFlight.flight_id == flight.id
    ).all()
    
    # Fill in flight information (top rows)
    # Row 2: "Дата:"
    ws["C2"] = flight.departure_date.strftime("%d.%m.%Y")
    
    # Row 3: "№ рейса:"
    ws["C3"] = flight.flight_number
    
    # Row 4: "Маршрут:"
    ws["C4"] = flight.route or ""
    
    # Row 5: "Время вылета:"
    departure_time = flight.departure_time.strftime('%H:%M') if flight.departure_time else ""
    ws["C5"] = departure_time
    
    # Passengers start from row 9
    # Headers are in row 8
    start_row = 9
    
    for idx, pf in enumerate(passenger_flights, 1):
        passenger = pf.passengers
        
        row = start_row + idx - 1
        
        # Column A: № (number)
        ws[f"A{row}"] = idx
        
        # Column B: Фамилия, Имя, Отчество
        ws[f"B{row}"] = passenger.fullname
        
        # Column C: Пол (Gender)
        ws[f"C{row}"] = passenger.gender.value if passenger.gender else ""
        
        # Column D: Дата рождения
        ws[f"D{row}"] = passenger.birthdate.strftime("%d.%m.%Y") if passenger.birthdate else ""
        
        # Column E: № документа (Passport)
        ws[f"E{row}"] = str(passenger.passport)
        
        # Column F: Класс (keep existing "Y")
        ws[f"F{row}"] = "Y"
        
        # Column G: Гражданство (Nationality)
        ws[f"G{row}"] = "РФ"  # Assuming Russian Federation
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_both_excel_files(flight: Flight, session: Session) -> tuple[bytes, bytes]:
    """
    Generate both Excel files for a flight.
    
    Returns:
        Tuple of (manifest_bytes, ticket_list_bytes)
    """
    manifest = generate_passenger_manifest_excel(flight, session)
    ticket_list = generate_ticket_issuance_list_excel(flight, session)
    return manifest, ticket_list